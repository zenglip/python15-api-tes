#name

from openpyxl import load_workbook
from API_0419.common import http_request
from API_0419.common import contants
import  json

class Case:
    '''测试类，每一个测试用例都是该类的一个实例对象'''
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql = None
class excel_practice:
    '''读写表格数据'''
    def __init__(self,excel_name,sheet_name):
        self.excel_name = excel_name
        self.sheet_name = sheet_name
        self.wb = load_workbook(self.excel_name)
        self.sheet = self.wb[self.sheet_name]


    def read_excel(self):
        '''读数据'''
        wb = self.wb
        sheet = self.sheet
        max_row = sheet.max_row#获取最大行
        max_L = sheet.max_column
        List_cases = []
        for i in range(2,max_row+1):
            case = Case()
            case.case_id = sheet.cell(i,1).value
            case.title = sheet.cell(i, 2).value
            case.url = sheet.cell(i, 3).value
            case.data = sheet.cell(i, 4).value
            case.method = sheet.cell(i, 5).value
            case.expected = sheet.cell(i, 6).value
            case.sql = sheet.cell(i,9).value
            List_cases.append(case)
        wb.close()
        return List_cases
    def write_excel(self,row,actual,result):
        '''写数据'''
        sheet = self.wb[self.sheet_name]
        sheet.cell(row,7,actual)
        sheet.cell(row,8,result)
        self.wb.save(self.excel_name)
        self.wb.close()

if __name__ == '__main__':
    excel = excel_practice('cases.xlsx', "invest")
    cases = excel.read_excel()

    #先实例化一个对象
    # http_request = http_request.Http_Request2()
    # #注册
    # # do_excel = excel_practice("cases.xlsx", sheet_name="register")
    # # cases = do_excel.read_excel()
    #
    # #先登录
    # # do_excel = excel_practice("cases.xlsx",sheet_name= "login")
    # # cases = do_excel.read_excel()
    #
    # #在充值
    # do_excel = excel_practice("cases.xlsx", "recharge")
    # cases = do_excel.read_excel()
    # #取现
    # # do_excel = excel_practice("cases.xlsx", "withdraw")
    # # cases = do_excel.read_excel()
    # #竞标
    # # do_excel = excel_practice("cases.xlsx", "bidLoan")
    # # cases = do_excel.read_excel()
    # #新增
    # # do_excel = excel_practice("cases.xlsx", "add")
    # # cases = do_excel.read_excel()
    #
    # for case in cases:
    #     print(case.__dict__)
    #     resp = http_request.request(url = case.url,method = case.method,data = case.data)
    #     #注意调用注册方法时，注意手机号是否已经被注册过，否则第一条用例会出现第一条应该通过的用例报出（手机号码已被注册）
    #     #resp_dict = resp.json()  # 返回字典
    #
    #     actual = resp.text
    #     actual = json.loads(actual)
    #     actual.pop('data')
    #
    #     #actual.pop("data")#data中包含余额，每次请求后余额会发生变化，所以将data删除后比较,也可以充值金额为0
    #     actual = str(actual)
    #     # print(actual)
    #     # print(case.expected)
    #     if case.expected == actual:#判断期望值和实际值是否相等
    #         do_excel.write_excel(case.case_id+1,actual,"PASS")
    #     else:
    #         do_excel.write_excel(case.case_id + 1, actual, "False")
    #
    # http_request.close()
    # ##使用unittest暂时还没有完成